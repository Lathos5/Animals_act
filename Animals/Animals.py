import wx
import wx.xrc

import os
import os.path

from openpyxl import load_workbook

from Models.Anim import Animal
from Models.Cats import Cat
from Models.Dogs import Dog
from Models.RegAnimal import signupWindow

class AnimalsFrame(wx.Frame):
    def __init__(self,parent,title):

        self.animals = ListAnimals()

        wx.Frame.__init__(self, None, title=title, size=(1080,720))
        self.archivo='untitled.txt'

        p=wx.Panel(self, -1)

        """
        Sizer
        """
        sz=wx.BoxSizer(wx.VERTICAL)

        """
        Editor
        """
        self.editor=wx.TextCtrl(p, -1, "", style=wx.TE_MULTILINE)

        """
        Adding to the sizer
        """
        sz.Add(self.editor, 1, wx.EXPAND)
        p.SetSizer(sz)

        """
        Creates menubar
        """
        self.crearMenu()
        self.Center(True)
        self.Show()



    def crearMenu(self):
        """
        def crearMenu(self):

        Crea la barra de menú 
        """
        mFile=wx.Menu()
        oFiles=mFile.Append(-1, "Open File")
        

        mAdd=wx.Menu()
        aAnimal=mAdd.Append(-1, "Add New Animal ")

        mPrint = wx.Menu()
        showAnimals = mPrint.Append(-1, "Show Animals")

        menuBar=wx.MenuBar()
        menuBar.Append(mFile, "File")
        menuBar.Append(mAdd, "Edit")
        menuBar.Append(mPrint, "Print Registers")

        self.SetMenuBar(menuBar)

        """
        Naming the "events"
        """
        self.Bind(wx.EVT_MENU, self.openFile, oFiles)
        
        self.Bind(wx.EVT_MENU, self.createAnimal, aAnimal)

        self.Bind(wx.EVT_MENU, self.printing, showAnimals)

 

    def form_cats(self, cname, cage, ccolor):
        """
        def form_cats(self, cname, cage, ccolor):

        Method used to get values from the cats Form
        """
        cat = Cat(name = cname,
                  color=ccolor,
                  age=cage)
        self.animals.lAnimals.append(cat)


    def form_dogs(self, dname, dage, dcolor):
        """
        def form_dogs(self, dname, dage, dcolor):

        Method used to get values from the dogs Form
        """
        dog = Dog(name = dname,
                  color=dcolor,
                  friendly = '', 
                  hasOwner = '',
                  age=dage)
        self.animals.lAnimals.append(dog)


    def openFile(self, event):
        """
        def openFile(self, event):

        Method used to open a file to attach an animal
        """
        wc = "Text Files (*.txt)|*.txt|Excel Files (*.xslx)|*.xlsx"
        dlg=wx.FileDialog(self, "Open Animals File", wildcard=wc, style = wx.FD_OPEN)
        if dlg.ShowModal() == wx.ID_CANCEL:
            return
        pathname = dlg.GetPath()

        filename, extension = os.path.splitext(pathname)

        """
        According to the file extension, we call the appropiate method for the file
        """

        if extension == ".txt":
            """
            method created for txt files
            """
            self.txtFile(pathname, dlg)
            
        elif extension == ".xlsx":
            """
            method created for excel files
            """
            self.xslxFile(pathname, dlg)

        
    def xslxFile(self, pathname, dialog):
        try:
            """
            def xslxFile(self, pathname, dialog):

            Opens a excel file format

            Dogs
            """
            workbook = load_workbook(filename = pathname)
            sheet = workbook.active

           
            for row in sheet.iter_rows(min_row = 2, values_only = True):
                """
                for each row on the excel dogs file, we will create a dog object whom is going to be added to the animals list
                """
                dog = Dog(name=row[0],
                         color=row[1],
                         friendly=row[2],
                         hasOwner=row[3],
                         age=row[4])
                self.animals.lAnimals.append(dog)
            self.alert("Dogs")
        except IOError:
             wx.LogError(u"The file is unavailable '%s'." % newfile)
             dialog.Destroy()
  
    
    
    def txtFile(self, pathname, dialog):
        try:
            with open(pathname) as Gatos:
                """
                def txtFile(self, pathname, dialog):

                Opens a txt file format

                Cats
                """
                for gato in Gatos.readlines():
                    """                   
                    Removing multiple spaces
                    """      
                    while '  ' in gato:
                        gato = gato.replace('\n', '')
                        gato = gato.replace('  ',' ')
                        arrayInfo = gato.split(' ')

                        """
                        excluding the header from the cats file to store on the animals list
                        """

                    if arrayInfo[0] != 'name':
                        cat = Cat(name = arrayInfo[0],
                                  color=arrayInfo[1],
                                  age=arrayInfo[2])
                        self.animals.lAnimals.append(cat)
            Gatos.close()
            self.alert("Cats")
        except IOError:
            wx.LogError(u"The file is unavailable '%s'." % newfile)
            dialog.Destroy()
  
    
    def alert(self, pet):
        """
        def alert(self, pet):

        Shows a message acorrding to the type of animal attached form the Open File option
        """
        wx.MessageBox(pet +' have been added', 'Info', wx.OK | wx.ICON_INFORMATION)
    
    def printing(self, event):
        """
        def printing(self, event):

        Method used to shhow the animals list
        """
        self.editor.SetValue(self.animals.shownList())

    def createAnimal(self, event):
        """
        def createAnimal(self, event):

        Opens the dogs frame
        """
        secondWindow = signupWindow(self)
        secondWindow.Show()



class ListAnimals():
    """
    Class for the animals list
    """
    lAnimals = []

    def shownList(self):
        """
        method used to create the output, concatenating a string by each tuple of the animals list
        """
        self.lAnimals = sorted(self.lAnimals, key=lambda x: x.name)
        result = ''
        for animal in self.lAnimals:
            result += str (animal) + "\n"
        return result

#"""
#Documentation of the entire proyect, including classes and methods
#"""

help(AnimalsFrame)
help(ListAnimals)
help(Dog)
help(Cat)
help(Animal)
help(signupWindow)

if __name__=='__main__':
    app = wx.App()
    fr = AnimalsFrame(None, "Animals")
    app.MainLoop()